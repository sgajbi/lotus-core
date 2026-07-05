from __future__ import annotations

import csv
from csv import Error as CsvError
from dataclasses import dataclass
from io import BytesIO, StringIO
from itertools import zip_longest
from typing import Any, Iterable, Literal
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import BaseModel, ValidationError

from ..application.errors import UnsupportedOperation, ValidationRejected
from ..application.upload_commands import UploadEntity, UploadRowIssue
from ..DTOs.business_date_dto import BusinessDate
from ..DTOs.fx_rate_dto import FxRate
from ..DTOs.instrument_dto import Instrument
from ..DTOs.market_price_dto import MarketPrice
from ..DTOs.portfolio_dto import Portfolio
from ..DTOs.transaction_dto import Transaction

MODEL_BY_ENTITY: dict[UploadEntity, type[BaseModel]] = {
    "portfolios": Portfolio,
    "instruments": Instrument,
    "transactions": Transaction,
    "market_prices": MarketPrice,
    "fx_rates": FxRate,
    "business_dates": BusinessDate,
}
DEFAULT_UPLOAD_MAX_ROWS = 5_000
DEFAULT_UPLOAD_MAX_COLUMNS = 200
DEFAULT_UPLOAD_MAX_CELL_LENGTH = 8_192
UPLOAD_PARSER_BUDGET_REASON_CODE = "upload_parser_budget_exceeded"
UPLOAD_PARSER_BUDGET_ERROR_CODE = "INGESTION_UPLOAD_PARSER_BUDGET_EXCEEDED"


@dataclass(frozen=True, slots=True)
class UploadParserBudget:
    max_rows: int = DEFAULT_UPLOAD_MAX_ROWS
    max_columns: int = DEFAULT_UPLOAD_MAX_COLUMNS
    max_cell_length: int = DEFAULT_UPLOAD_MAX_CELL_LENGTH


@dataclass(frozen=True, slots=True)
class UploadValidationReport:
    file_format: Literal["csv", "xlsx"]
    valid_models: list[BaseModel]
    valid_rows: list[dict[str, Any]]
    errors: list[UploadRowIssue]
    total_rows: int


class BulkUploadValidator:
    def __init__(self, *, budget: UploadParserBudget | None = None) -> None:
        self._budget = budget or UploadParserBudget()

    def validate(
        self,
        *,
        entity_type: UploadEntity,
        filename: str,
        content: bytes,
    ) -> UploadValidationReport:
        model_cls = MODEL_BY_ENTITY[entity_type]
        alias_index = _field_alias_index(model_cls)
        file_format, rows = _parse_rows(filename, content, self._budget)

        valid_models: list[BaseModel] = []
        valid_rows: list[dict[str, Any]] = []
        errors: list[UploadRowIssue] = []

        for index, row in enumerate(rows, start=2):
            normalized_row = _normalize_row(row, alias_index)
            try:
                model = model_cls.model_validate(normalized_row)
            except ValidationError as exc:
                errors.append(_row_issue(row_number=index, exc=exc))
                continue

            valid_models.append(model)
            valid_rows.append(model.model_dump())

        return UploadValidationReport(
            file_format=file_format,
            valid_models=valid_models,
            valid_rows=valid_rows,
            errors=errors,
            total_rows=len(rows),
        )


def _row_issue(*, row_number: int, exc: ValidationError) -> UploadRowIssue:
    issues: list[str] = []
    for error in exc.errors():
        location = ".".join(str(part) for part in error.get("loc", ()))
        issues.append(f"{location}: {error.get('msg', 'invalid value')}")
    return UploadRowIssue(row_number=row_number, message="; ".join(issues))


def _normalized_key(value: str) -> str:
    return "".join(ch for ch in value.lower() if ch.isalnum())


def _field_alias_index(model_cls: type[BaseModel]) -> dict[str, str]:
    index: dict[str, str] = {}
    for field_name, field_info in model_cls.model_fields.items():
        index[_normalized_key(field_name)] = field_name
        alias = field_info.alias
        if alias:
            index[_normalized_key(alias)] = alias
    return index


def _normalize_row(row: dict[str, Any], alias_index: dict[str, str]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for raw_key, raw_value in row.items():
        canonical_key = _canonical_row_key(raw_key, alias_index)
        if canonical_key is None:
            continue
        normalized[canonical_key] = _normalized_row_value(raw_value)
    return normalized


def _canonical_row_key(raw_key: Any, alias_index: dict[str, str]) -> str | None:
    if raw_key is None:
        return None
    return alias_index.get(_normalized_key(str(raw_key)))


def _normalized_row_value(raw_value: Any) -> Any:
    if not isinstance(raw_value, str):
        return raw_value
    stripped = raw_value.strip()
    return None if stripped == "" else stripped


def _parse_rows(
    filename: str,
    content: bytes,
    budget: UploadParserBudget,
) -> tuple[Literal["csv", "xlsx"], list[dict[str, Any]]]:
    file_format = _detect_format(filename)
    try:
        if file_format == "csv":
            return file_format, _parse_csv(content, budget)
        return file_format, _parse_xlsx(content, budget)
    except (UnicodeDecodeError, CsvError) as exc:
        raise ValidationRejected(
            reason_code="invalid_csv_content",
            detail=f"Invalid CSV content: {exc}",
        ) from exc
    except (BadZipFile, InvalidFileException, ValueError) as exc:
        raise ValidationRejected(
            reason_code="invalid_xlsx_content",
            detail=f"Invalid XLSX content: {exc}",
        ) from exc


def _parse_csv(content: bytes, budget: UploadParserBudget) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig")
    reader = csv.reader(StringIO(text))
    headers = next(reader, None)
    if headers is None:
        return []
    _validate_row_shape(
        row_values=headers,
        row_number=1,
        budget=budget,
        row_kind="header",
    )

    records: list[dict[str, Any]] = []
    for row_number, row_values in enumerate(reader, start=2):
        _validate_row_budget(records, budget)
        _validate_row_shape(
            row_values=row_values,
            row_number=row_number,
            budget=budget,
            row_kind="row",
        )
        records.append(_csv_row_record(headers, row_values))
    return records


def _parse_xlsx(content: bytes, budget: UploadParserBudget) -> list[dict[str, Any]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            return []

        _validate_row_shape(
            row_values=header_row,
            row_number=1,
            budget=budget,
            row_kind="header",
        )
        headers = _xlsx_headers(header_row)
        records: list[dict[str, Any]] = []
        for row_number, row_values in enumerate(rows, start=2):
            _validate_row_shape(
                row_values=row_values,
                row_number=row_number,
                budget=budget,
                row_kind="row",
            )
            row_dict = _xlsx_row_record(headers, row_values)
            if _xlsx_row_has_data(row_dict):
                _validate_row_budget(records, budget)
                records.append(row_dict)
        return records
    finally:
        workbook.close()


def _csv_row_record(headers: list[str], row_values: list[str]) -> dict[str, Any]:
    return {
        header: value
        for header, value in zip_longest(headers, row_values, fillvalue=None)
        if header
    }


def _validate_row_budget(records: list[dict[str, Any]], budget: UploadParserBudget) -> None:
    if len(records) >= budget.max_rows:
        _raise_parser_budget_exceeded(
            message="Bulk upload row count exceeds the configured parser budget.",
            budget_name="max_rows",
            limit=budget.max_rows,
        )


def _validate_row_shape(
    *,
    row_values: Iterable[Any],
    row_number: int,
    budget: UploadParserBudget,
    row_kind: str,
) -> None:
    values = tuple(row_values)
    if len(values) > budget.max_columns:
        _raise_parser_budget_exceeded(
            message="Bulk upload column count exceeds the configured parser budget.",
            budget_name="max_columns",
            limit=budget.max_columns,
            observed=len(values),
            row_number=row_number,
        )
    for column_index, value in enumerate(values, start=1):
        if value is not None and len(str(value)) > budget.max_cell_length:
            _raise_parser_budget_exceeded(
                message="Bulk upload cell length exceeds the configured parser budget.",
                budget_name="max_cell_length",
                limit=budget.max_cell_length,
                observed=len(str(value)),
                row_number=row_number,
                column_index=column_index,
                row_kind=row_kind,
            )


def _raise_parser_budget_exceeded(
    *,
    message: str,
    budget_name: str,
    limit: int,
    observed: int | None = None,
    row_number: int | None = None,
    column_index: int | None = None,
    row_kind: str | None = None,
) -> None:
    detail: dict[str, Any] = {
        "code": UPLOAD_PARSER_BUDGET_ERROR_CODE,
        "message": message,
        "budget": budget_name,
        "limit": limit,
    }
    if observed is not None:
        detail["observed"] = observed
    if row_number is not None:
        detail["row_number"] = row_number
    if column_index is not None:
        detail["column_index"] = column_index
    if row_kind is not None:
        detail["row_kind"] = row_kind
    raise ValidationRejected(
        reason_code=UPLOAD_PARSER_BUDGET_REASON_CODE,
        detail=detail,
    )


def _xlsx_headers(row: tuple[Any, ...]) -> list[str]:
    return [str(cell).strip() if cell is not None else "" for cell in row]


def _xlsx_row_record(headers: list[str], row_values: tuple[Any, ...] | None) -> dict[str, Any]:
    if row_values is None:
        return {}
    return {
        header: row_values[index] if index < len(row_values) else None
        for index, header in enumerate(headers)
        if header
    }


def _xlsx_row_has_data(row: dict[str, Any]) -> bool:
    return any(value is not None and str(value).strip() != "" for value in row.values())


def _detect_format(filename: str) -> Literal["csv", "xlsx"]:
    lowered = filename.lower()
    if lowered.endswith(".csv"):
        return "csv"
    if lowered.endswith(".xlsx"):
        return "xlsx"
    raise UnsupportedOperation(
        reason_code="unsupported_upload_file_format",
        detail="Unsupported file format. Use .csv or .xlsx.",
    )
