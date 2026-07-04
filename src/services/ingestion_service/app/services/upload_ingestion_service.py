from __future__ import annotations

import csv
from csv import Error as CsvError
from dataclasses import dataclass
from io import BytesIO, StringIO
from typing import Any, Literal
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import BaseModel, ValidationError

from ..application.errors import UnsupportedOperation, ValidationRejected
from ..application.upload_commands import (
    UploadCommitCommand,
    UploadCommitResult,
    UploadEntity,
    UploadPreviewCommand,
    UploadPreviewResult,
    UploadRowIssue,
)
from ..DTOs.business_date_dto import BusinessDate
from ..DTOs.fx_rate_dto import FxRate
from ..DTOs.instrument_dto import Instrument
from ..DTOs.market_price_dto import MarketPrice
from ..DTOs.portfolio_dto import Portfolio
from ..DTOs.transaction_dto import Transaction
from ..services.ingestion_service import IngestionService

MODEL_BY_ENTITY: dict[UploadEntity, type[BaseModel]] = {
    "portfolios": Portfolio,
    "instruments": Instrument,
    "transactions": Transaction,
    "market_prices": MarketPrice,
    "fx_rates": FxRate,
    "business_dates": BusinessDate,
}


def _normalized_key(value: str) -> str:
    return "".join(ch for ch in value.lower() if ch.isalnum())


def _field_alias_index(model_cls: type[BaseModel]) -> dict[str, str]:
    index: dict[str, str] = {}
    for field_name, field_info in model_cls.model_fields.items():
        index[_normalized_key(field_name)] = field_name
        alias = field_info.alias
        if alias:
            index[_normalized_key(alias)] = alias
    return index


def _normalize_row(row: dict[str, Any], alias_index: dict[str, str]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for raw_key, raw_value in row.items():
        canonical_key = _canonical_row_key(raw_key, alias_index)
        if canonical_key is None:
            continue
        normalized[canonical_key] = _normalized_row_value(raw_value)
    return normalized


def _canonical_row_key(raw_key: Any, alias_index: dict[str, str]) -> str | None:
    if raw_key is None:
        return None
    return alias_index.get(_normalized_key(str(raw_key)))


def _normalized_row_value(raw_value: Any) -> Any:
    if not isinstance(raw_value, str):
        return raw_value
    stripped = raw_value.strip()
    return None if stripped == "" else stripped


def _parse_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))
    return [dict(row) for row in reader]


def _parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = _xlsx_headers(rows[0])
    records: list[dict[str, Any]] = []
    for row_values in rows[1:]:
        row_dict = _xlsx_row_record(headers, row_values)
        if _xlsx_row_has_data(row_dict):
            records.append(row_dict)
    return records


def _xlsx_headers(row: tuple[Any, ...]) -> list[str]:
    return [str(cell).strip() if cell is not None else "" for cell in row]


def _xlsx_row_record(headers: list[str], row_values: tuple[Any, ...] | None) -> dict[str, Any]:
    if row_values is None:
        return {}
    return {
        header: row_values[index] if index < len(row_values) else None
        for index, header in enumerate(headers)
        if header
    }


def _xlsx_row_has_data(row: dict[str, Any]) -> bool:
    return any(value is not None and str(value).strip() != "" for value in row.values())


def _detect_format(filename: str) -> Literal["csv", "xlsx"]:
    lowered = filename.lower()
    if lowered.endswith(".csv"):
        return "csv"
    if lowered.endswith(".xlsx"):
        return "xlsx"
    raise UnsupportedOperation(
        reason_code="unsupported_upload_file_format",
        detail="Unsupported file format. Use .csv or .xlsx.",
    )


@dataclass
class _ValidationResult:
    file_format: Literal["csv", "xlsx"]
    valid_models: list[BaseModel]
    valid_rows: list[dict[str, Any]]
    errors: list[UploadRowIssue]
    total_rows: int


class UploadIngestionService:
    def __init__(self, ingestion_service: IngestionService):
        self._ingestion_service = ingestion_service

    def _parse_rows(
        self, filename: str, content: bytes
    ) -> tuple[Literal["csv", "xlsx"], list[dict[str, Any]]]:
        file_format = _detect_format(filename)
        try:
            if file_format == "csv":
                return file_format, _parse_csv(content)
            return file_format, _parse_xlsx(content)
        except (UnicodeDecodeError, CsvError) as exc:
            raise ValidationRejected(
                reason_code="invalid_csv_content",
                detail=f"Invalid CSV content: {exc}",
            ) from exc
        except (BadZipFile, InvalidFileException, ValueError) as exc:
            raise ValidationRejected(
                reason_code="invalid_xlsx_content",
                detail=f"Invalid XLSX content: {exc}",
            ) from exc

    def _validate_rows(
        self, entity_type: UploadEntity, filename: str, content: bytes
    ) -> _ValidationResult:
        model_cls = MODEL_BY_ENTITY[entity_type]
        alias_index = _field_alias_index(model_cls)
        file_format, rows = self._parse_rows(filename, content)

        valid_models: list[BaseModel] = []
        valid_rows: list[dict[str, Any]] = []
        errors: list[UploadRowIssue] = []

        for index, row in enumerate(rows, start=2):
            normalized_row = _normalize_row(row, alias_index)
            try:
                model = model_cls.model_validate(normalized_row)
            except ValidationError as exc:
                issues: list[str] = []
                for error in exc.errors():
                    location = ".".join(str(part) for part in error.get("loc", ()))
                    issues.append(f"{location}: {error.get('msg', 'invalid value')}")
                errors.append(UploadRowIssue(row_number=index, message="; ".join(issues)))
                continue

            valid_models.append(model)
            valid_rows.append(model.model_dump())

        return _ValidationResult(
            file_format=file_format,
            valid_models=valid_models,
            valid_rows=valid_rows,
            errors=errors,
            total_rows=len(rows),
        )

    def preview_upload(
        self,
        command: UploadPreviewCommand,
    ) -> UploadPreviewResult:
        validation = self._validate_rows(
            command.entity_type,
            command.filename,
            command.content,
        )
        return UploadPreviewResult(
            entity_type=command.entity_type,
            file_format=validation.file_format,
            total_rows=validation.total_rows,
            valid_rows=len(validation.valid_models),
            invalid_rows=len(validation.errors),
            sample_rows=validation.valid_rows[: command.sample_size],
            errors=validation.errors[: command.sample_size],
        )

    async def commit_upload(
        self,
        command: UploadCommitCommand,
    ) -> UploadCommitResult:
        validation = self._validate_rows(
            command.entity_type,
            command.filename,
            command.content,
        )
        self._validate_commit(validation, command.allow_partial)
        await self._publish_valid_models(command.entity_type, validation.valid_models)
        return self._commit_response(command.entity_type, validation)

    def _validate_commit(self, validation: _ValidationResult, allow_partial: bool) -> None:
        if validation.total_rows == 0:
            self._raise_empty_upload()
        if validation.errors and not allow_partial:
            self._raise_partial_upload_rejected(validation.errors)
        if not validation.valid_models:
            self._raise_no_valid_upload_rows()

    def _raise_empty_upload(self) -> None:
        raise ValidationRejected(
            reason_code="empty_upload",
            detail="Upload file contains no data rows.",
        )

    def _raise_partial_upload_rejected(self, errors: list[UploadRowIssue]) -> None:
        raise ValidationRejected(
            reason_code="upload_invalid_rows",
            detail={
                "message": "Upload contains invalid rows. Fix errors or use allow_partial=true.",
                "errors": [error.as_dict() for error in errors[:50]],
            },
        )

    def _raise_no_valid_upload_rows(self) -> None:
        raise ValidationRejected(
            reason_code="upload_no_valid_rows",
            detail="No valid rows found in upload.",
        )

    async def _publish_valid_models(
        self, entity_type: UploadEntity, valid_models: list[BaseModel]
    ) -> None:
        publishers = {
            "portfolios": self._publish_portfolios,
            "instruments": self._publish_instruments,
            "transactions": self._publish_transactions,
            "market_prices": self._publish_market_prices,
            "fx_rates": self._publish_fx_rates,
            "business_dates": self._publish_business_dates,
        }
        await publishers[entity_type](valid_models)

    async def _publish_portfolios(self, valid_models: list[BaseModel]) -> None:
        await self._ingestion_service.publish_portfolios(
            [model for model in valid_models if isinstance(model, Portfolio)]
        )

    async def _publish_instruments(self, valid_models: list[BaseModel]) -> None:
        await self._ingestion_service.publish_instruments(
            [model for model in valid_models if isinstance(model, Instrument)]
        )

    async def _publish_transactions(self, valid_models: list[BaseModel]) -> None:
        await self._ingestion_service.publish_transactions(
            [model for model in valid_models if isinstance(model, Transaction)]
        )

    async def _publish_market_prices(self, valid_models: list[BaseModel]) -> None:
        await self._ingestion_service.publish_market_prices(
            [model for model in valid_models if isinstance(model, MarketPrice)]
        )

    async def _publish_fx_rates(self, valid_models: list[BaseModel]) -> None:
        await self._ingestion_service.publish_fx_rates(
            [model for model in valid_models if isinstance(model, FxRate)]
        )

    async def _publish_business_dates(self, valid_models: list[BaseModel]) -> None:
        await self._ingestion_service.publish_business_dates(
            [model for model in valid_models if isinstance(model, BusinessDate)]
        )

    def _commit_response(
        self, entity_type: UploadEntity, validation: _ValidationResult
    ) -> UploadCommitResult:
        return UploadCommitResult(
            entity_type=entity_type,
            file_format=validation.file_format,
            total_rows=validation.total_rows,
            valid_rows=len(validation.valid_models),
            invalid_rows=len(validation.errors),
            published_rows=len(validation.valid_models),
            skipped_rows=len(validation.errors),
            message="Upload committed and queued for processing.",
        )
